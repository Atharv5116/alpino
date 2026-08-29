"""Bulk importer for Buyer Master sheets exported from the Data Import tool.

Run preflight() first, then run(). run() aborts while blocking issues remain and prints the flag
that clears each one.
"""

import csv
import os
import re
from collections import Counter, OrderedDict, defaultdict

import frappe

DOCTYPE = "Buyer Master"

# CSV label -> (child table or "" for Buyer Master itself, fieldname); fieldname None = read but not imported
COLUMN_MAP = OrderedDict(
	[
		("ID", ("", "name")),
		("Customer Type", ("", "customer_type")),
		("Customer (Business Name)", ("", "customer_business_name")),
		("GST Type", ("", "gst_type")),
		("Level", ("", "level")),
		("Payment Term", ("", "payment_term")),
		("Email", ("", "email")),
		("Contact No", ("", "contact_no")),
		("Contact Person", ("", "contact_person")),
		("Channel", ("", "channel")),
		("Appointment Required", ("", "appointment_required")),
		("GRN Available", ("", "grn_available")),
		("Partial Order Allowed", ("", "partial_order_allowed")),
		("GST-Exclusive Buyer", ("", "gst_exclusive_buyer")),
		("Site Name", ("", "site_name")),
		("Is Parent", ("", "is_parent")),
		("Parent Buyer", ("", "parent_buyer")),
		("Parent Business Name", ("", "parent_business_name")),
		("Customer ID", ("", "customer_id")),
		("GST No", ("", "gst_no")),
		("GST Certificate", ("", "gst_certificate")),
		("PAN No", ("", "pan_no")),
		("PAN Attachment", ("", "pan_attachment")),
		("Days", ("", "payment_term_days")),
		("Party Owner", ("", "party_owner")),
		("Tally Buyer Name", ("", "tally_buyer_name")),
		("Tally P&L Name", ("", "tally_pl_name")),
		("Tally Warehouse Id", ("", "custom_tally_warehouse_id")),
		("Combine Product Bundles", ("", "combine_product_bundles")),
		("Primary POC", ("", "primary_poc")),
		("Secondary POC", ("", "secondary_poc")),
		("Employee Name", ("", "poc_employee")),
		("Shipping Same as Primary", ("", "shipping_same_as_profile")),
		("Shipping Address", ("", "shipping_address")),
		("Shipping State", ("", "shipping_state")),
		("Shipping City", ("", "shipping_city")),
		("Customer", ("", "customer")),
		("Address (primary sync)", ("", "address")),
		("Pincode (primary sync)", ("", "pincode")),
		("Country (primary sync)", ("", "country")),
		("State (primary sync)", ("", "state")),
		("City (primary sync)", ("", "city")),
		("Area (primary sync)", ("", "area")),
		("Sub Area (primary sync)", ("", "sub_area")),
		("Alternate No", ("", "alternate_no")),
		# --- Addresses child table ---
		("ID (Addresses)", ("addresses", None)),  # source row name, not reusable
		("Address (Addresses)", ("addresses", "address_line")),
		("Area (Addresses)", ("addresses", "area")),
		("City (Addresses)", ("addresses", "city")),
		("Country (Addresses)", ("addresses", "country")),
		("Label (Addresses)", ("addresses", "address_label")),
		("Pincode (Addresses)", ("addresses", "pincode")),
		("Primary (Addresses)", ("addresses", "is_primary")),
		("Shipping (Addresses)", ("addresses", "is_shipping")),
		("Site Name (Addresses)", ("addresses", "site_name")),
		("State (Addresses)", ("addresses", "state")),
		("Sub Area (Addresses)", ("addresses", "sub_area")),
		# --- Margins child table ---
		("ID (Margins)", ("margins", None)),
		("Item Group (Margins)", ("margins", "item_group")),
		("Margin % (Margins)", ("margins", "margin_percent")),
		("Product Name (Margins)", ("margins", "product_name")),
		("SKU (Margins)", ("margins", "sku")),
	]
)

CHECK_FIELDS = {
	"appointment_required",
	"grn_available",
	"partial_order_allowed",
	"gst_exclusive_buyer",
	"is_parent",
	"combine_product_bundles",
	"shipping_same_as_profile",
	"is_primary",
	"is_shipping",
}

TRUE_VALUES = {"1", "1.0", "y", "yes", "true", "checked"}


def _cell(fieldname, val):
	"""Sheet cell to python value; force checkboxes to a real 0/1 since the string "0" is truthy."""
	if fieldname in CHECK_FIELDS:
		return 1 if str(val).strip().lower() in TRUE_VALUES else 0
	return val


MANDATORY = (
	"customer_business_name",
	"customer_type",
	"gst_type",
	"level",
	"payment_term",
	"email",
	"contact_no",
	"contact_person",
)

# Link fields sanitised before saving: fieldname -> target doctype
PARENT_LINKS = {
	"customer_type": "Alpino Customer Type",
	"channel": "Channel",
	"party_owner": "User",
	"poc_employee": "Employee",
	"country": "Country",
	"state": "State",
	"city": "City",
	"shipping_state": "State",
	"shipping_city": "City",
}
ADDRESS_LINKS = {"city": "City", "state": "State", "country": "Country"}
MARGIN_LINKS = {"item_group": "Item Group", "sku": "Item"}

GSTIN_RE = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$")
PIN_RE = re.compile(r"^[1-9][0-9]{5}$")

# issue category -> (severity, what resolves it)
SEVERITY = {
	"customer-to-create": ("auto", "created with the sheet's own Customer ID"),
	"missing-sku": ("auto", "on_missing_sku"),
	"missing-item-group": ("auto", "cleared"),
	"missing-state": ("blocking", "create_masters=True"),
	"missing-city": ("blocking", "create_masters=True"),
	"missing-link": ("blocking", "clear_missing_links=True"),
	"duplicate-gstin": ("blocking", "allow_duplicate_gst=True"),
	"duplicate-customer": ("blocking", "needs the per-site customer-ID derivation in buyer_master.py"),
	"customer-taken": ("blocking", "needs the per-site customer-ID derivation in buyer_master.py"),
	"id-collision": ("blocking", "renumber_collisions=True to import them under a fresh ID"),
	"id-collision-parent": ("blocking", "fix the sheet — the ID is a Parent Buyer, so it cannot be renumbered"),
	"parent-not-found": ("blocking", "fix the sheet"),
	"parent-and-is-parent": ("blocking", "fix the sheet"),
	"mandatory": ("blocking", "fix the sheet"),
	"payment-days": ("blocking", "fix the sheet"),
	"gstin-format": ("blocking", "fix the sheet"),
	"pincode": ("blocking", "fix the sheet"),
	"primary-address": ("blocking", "fix the sheet"),
	"no-address": ("blocking", "fix the sheet"),
	"margin-percent": ("blocking", "fix the sheet"),
	"multiple-primary": ("blocking", "fix the sheet"),
}


# ---------------------------------------------------------------- parsing


def _read_rows(path):
	"""Return (header, rows) from a csv or xlsx export."""
	if path.lower().endswith((".xlsx", ".xls")):
		from openpyxl import load_workbook

		wb = load_workbook(path, read_only=True, data_only=True)
		ws = wb[wb.sheetnames[0]]
		rows = [["" if c is None else str(c).strip() for c in r] for r in ws.iter_rows(values_only=True)]
		wb.close()
	else:
		csv.field_size_limit(10**9)
		with open(path, newline="", encoding="utf-8-sig") as f:
			rows = [[(c or "").strip() for c in r] for r in csv.reader(f)]
	rows = [r for r in rows if any(r)]
	if not rows:
		frappe.throw(f"{path} is empty")
	return rows[0], rows[1:]


def _child_payload(row, idx, table):
	"""Pull one child-table row out of a sheet row; None when every cell is blank."""
	out = {}
	for label, (target, fieldname) in COLUMN_MAP.items():
		if target != table or not fieldname or label not in idx:
			continue
		val = row[idx[label]] if idx[label] < len(row) else ""
		if val == "":
			continue
		out[fieldname] = _cell(fieldname, val)
	return out or None


def parse(path):
	"""Sheet to a list of record dicts, in file order. An ID row starts a record; blank-ID rows add child rows to it."""
	header, rows = _read_rows(path)
	idx = {h: i for i, h in enumerate(header)}
	unknown = [h for h in header if h and h not in COLUMN_MAP]
	if unknown:
		frappe.throw(f"Unmapped column(s) in {os.path.basename(path)}: {unknown}")
	if "ID" not in idx:
		frappe.throw("The sheet needs an ID column (export it from Data Import with 'Update existing records').")

	records, cur = [], None
	for line_no, row in enumerate(rows, start=2):
		name = row[idx["ID"]] if idx["ID"] < len(row) else ""
		if name:
			cur = {"name": name, "line": line_no, "doc": {}, "addresses": [], "margins": []}
			records.append(cur)
			for label, (target, fieldname) in COLUMN_MAP.items():
				if target or not fieldname or fieldname == "name" or label not in idx:
					continue
				val = row[idx[label]] if idx[label] < len(row) else ""
				if val == "":
					continue
				cur["doc"][fieldname] = _cell(fieldname, val)
		elif cur is None:
			continue  # stray rows before the first record
		for table in ("addresses", "margins"):
			child = _child_payload(row, idx, table)
			if child:
				cur[table].append(child)
	return records


def _order_parents_first(records):
	"""Parents before children, so parent_buyer resolves during insert."""
	by_name = {r["name"]: r for r in records}
	emitted, ordered = set(), []

	def emit(rec, seen):
		if rec["name"] in emitted:
			return
		parent = (rec["doc"].get("parent_buyer") or "").strip()
		if parent and parent in by_name and parent not in emitted and parent not in seen:
			emit(by_name[parent], seen | {rec["name"]})
		emitted.add(rec["name"])
		ordered.append(rec)

	for rec in records:
		emit(rec, {rec["name"]})
	return ordered


# ---------------------------------------------------------------- analysis


def _existing(doctype, values):
	values = {v for v in values if v}
	if not values:
		return set()
	found, vals = set(), list(values)
	for i in range(0, len(vals), 500):  # keep the IN() clause sane
		found |= set(frappe.get_all(doctype, filters={"name": ["in", vals[i : i + 500]]}, pluck="name"))
	return found


def _lower(names):
	"""Lowercase names; MariaDB matches link values case-insensitively."""
	return {n.lower() for n in names}


def _has_derivation_fix():
	"""Whether this site's Buyer Master derives a distinct Customer per site."""
	from alpinos.alpinos_development.doctype.buyer_master import buyer_master

	return hasattr(buyer_master, "_customer_id_for_obm")


def _family_root(rec):
	"""The buyer family a record belongs to; a shared GSTIN is legal inside one."""
	d = rec["doc"]
	return (d.get("parent_buyer") or "").strip() or (rec["name"] if d.get("is_parent") else None)


def analyse(records):
	"""Resolve every link against the DB and collect issues."""
	wanted = defaultdict(set)
	for rec in records:
		for fieldname, dt in PARENT_LINKS.items():
			wanted[dt].add((rec["doc"].get(fieldname) or "").strip())
		wanted["Customer"].add((rec["doc"].get("customer") or "").strip())
		for a in rec["addresses"]:
			for fieldname, dt in ADDRESS_LINKS.items():
				wanted[dt].add((a.get(fieldname) or "").strip())
		for m in rec["margins"]:
			for fieldname, dt in MARGIN_LINKS.items():
				wanted[dt].add((m.get(fieldname) or "").strip())
	resolved = {dt: _existing(dt, vals) for dt, vals in wanted.items()}
	resolved[DOCTYPE] = _existing(DOCTYPE, {r["name"] for r in records})
	have = {dt: _lower(names) for dt, names in resolved.items()}

	# Customers already claimed by a Buyer Master other than the sheet's own row
	claimed = {}
	sheet_customers = wanted["Customer"] - {""}
	if sheet_customers:
		for row in frappe.get_all(
			DOCTYPE, filters={"customer": ["in", list(sheet_customers)]}, fields=["name", "customer"]
		):
			claimed[(row.customer or "").lower()] = row.name

	# sheet IDs that already name a different buyer here; updating one blends the two
	occupied, existing_gst = {}, {}
	if resolved[DOCTYPE]:
		for row in frappe.get_all(
			DOCTYPE,
			filters={"name": ["in", list(resolved[DOCTYPE])]},
			fields=["name", "customer_business_name", "gst_no"],
		):
			occupied[row.name] = (row.customer_business_name or "").strip()
			existing_gst[row.name] = (row.gst_no or "").strip().upper()

	# IDs another row depends on cannot be renumbered
	referenced_parents = {(r["doc"].get("parent_buyer") or "").strip() for r in records}

	# seed the GSTIN owner map from the DB too, so a new row reusing an existing
	# GSTIN trips the same "one GSTIN per buyer" check
	sheet_gsts = {
		(r["doc"].get("gst_no") or "").strip().upper()
		for r in records
		if r["doc"].get("gst_type") == "Registered Business"
	} - {""}
	gst_owner = {}
	if sheet_gsts:
		for row in frappe.get_all(
			DOCTYPE,
			filters={"gst_no": ["in", list(sheet_gsts)]},
			fields=["name", "gst_no", "parent_buyer", "is_parent"],
		):
			key = (row.gst_no or "").strip().upper()
			if key not in gst_owner:
				gst_owner[key] = (row.name, row.parent_buyer or (row.name if row.is_parent else None))

	names = {r["name"] for r in records}
	issues, gst_dupes, cust_conflicts = [], set(), set()
	# with the per-site derivation, a contested Customer can just be dropped and re-derived
	derivable = _has_derivation_fix()

	def add(rec, category, message, value="", severity=None):
		issues.append(
			{
				"name": rec["name"],
				"line": rec["line"],
				"category": category,
				"severity": severity or SEVERITY.get(category, ("blocking", ""))[0],
				"message": message,
				"value": value,
			}
		)

	customer_owner = {}

	for rec in records:
		d = rec["doc"]
		parent = (d.get("parent_buyer") or "").strip()

		here = occupied.get(rec["name"])
		sheet_biz = (d.get("customer_business_name") or "").strip()
		if here is not None and here.lower() != sheet_biz.lower():
			add(
				rec,
				"id-collision-parent" if rec["name"] in referenced_parents else "id-collision",
				f"{rec['name']} already exists on this site as '{here}', not '{sheet_biz}'",
				rec["name"],
			)

		for fieldname, dt in PARENT_LINKS.items():
			val = (d.get(fieldname) or "").strip()
			if not val or val.lower() in have[dt]:
				continue
			cat = {"State": "missing-state", "City": "missing-city"}.get(dt, "missing-link")
			add(rec, cat, f"{fieldname}: {dt} '{val}' does not exist", val)

		cust = (d.get("customer") or "").strip()
		if cust:
			key = cust.lower()
			if key not in have["Customer"]:
				add(rec, "customer-to-create", f"Customer '{cust}' will be created", cust)
			prev = customer_owner.get(key)
			if prev:
				cust_conflicts.add(rec["name"])
				add(
					rec,
					"duplicate-customer",
					f"shares Customer '{cust}' with {prev}"
					+ (" — the controller will derive a site-scoped ID" if derivable else " — one Buyer Master per Customer"),
					cust,
					severity="auto" if derivable else "blocking",
				)
			else:
				customer_owner[key] = rec["name"]
			if claimed.get(key) and claimed[key] != rec["name"]:
				cust_conflicts.add(rec["name"])
				add(
					rec,
					"customer-taken",
					f"Customer '{cust}' is already linked to {claimed[key]} in this site",
					cust,
					severity="auto" if derivable else "blocking",
				)

		if parent and parent not in names and parent.lower() not in have[DOCTYPE]:
			add(rec, "parent-not-found", f"parent_buyer '{parent}' is in neither the file nor the DB", parent)
		if parent and d.get("is_parent"):
			add(rec, "parent-and-is-parent", "Is Parent and Parent Buyer are both set")

		for fieldname in MANDATORY:
			if not str(d.get(fieldname) or "").strip():
				add(rec, "mandatory", f"missing mandatory {fieldname}")

		if d.get("payment_term") in ("Credit", "Partial") and not str(d.get("payment_term_days") or "").strip():
			add(rec, "payment-days", "Days is required when Payment Term is Credit/Partial")

		gst = (d.get("gst_no") or "").strip().upper()
		if gst and not GSTIN_RE.match(gst):
			add(rec, "gstin-format", f"invalid GSTIN format '{gst}'", gst)
		if gst and d.get("gst_type") == "Registered Business":
			root, prev = _family_root(rec), gst_owner.get(gst)
			if prev and prev[0] == rec["name"]:
				prev = None  # that holder is this very buyer, already in the DB
			if prev and not (root and prev[1] and root == prev[1]):
				# validate() only re-checks a GST that changed, so a buyer already
				# carrying this exact GSTIN here imports without tripping the rule.
				if existing_gst.get(rec["name"]) == gst:
					add(
						rec,
						"duplicate-gstin",
						f"GSTIN {gst} is also on {prev[0]}, an unrelated buyer — already recorded here, so it is not re-checked",
						gst,
						severity="auto",
					)
				else:
					add(rec, "duplicate-gstin", f"GSTIN {gst} is also on {prev[0]}, an unrelated buyer", gst)
					gst_dupes.add(rec["name"])
			else:
				gst_owner[gst] = (rec["name"], root)

		hierarchy = bool(d.get("is_parent") or parent)
		if not rec["addresses"]:
			if not hierarchy:
				add(rec, "no-address", "no address rows (required unless the buyer is a parent or a child site)")
		else:
			prim = [a for a in rec["addresses"] if a.get("is_primary")]
			if len(prim) > 1:
				add(rec, "multiple-primary", f"{len(prim)} address rows marked Primary — only one allowed")
			row = (prim or rec["addresses"])[0]
			if not hierarchy:
				for fieldname in ("address_line", "pincode", "country", "state", "city"):
					if not str(row.get(fieldname) or "").strip():
						add(rec, "primary-address", f"primary address row is missing {fieldname}")
			for a in rec["addresses"]:
				for fieldname, dt in ADDRESS_LINKS.items():
					val = (a.get(fieldname) or "").strip()
					if val and val.lower() not in have[dt]:
						cat = {"State": "missing-state", "City": "missing-city"}.get(dt, "missing-link")
						add(rec, cat, f"address {fieldname}: {dt} '{val}' does not exist", val)
				pin = (a.get("pincode") or "").strip()
				if pin and not PIN_RE.match(pin):
					add(rec, "pincode", f"invalid PIN '{pin}' — must be 6 digits, not starting with 0", pin)

		for m in rec["margins"]:
			sku = (m.get("sku") or "").strip()
			if sku and sku.lower() not in have["Item"]:
				add(rec, "missing-sku", f"margin SKU '{sku}' is not an Item", sku)
			ig = (m.get("item_group") or "").strip()
			if ig and ig.lower() not in have["Item Group"]:
				add(rec, "missing-item-group", f"margin Item Group '{ig}' does not exist", ig)
			if not str(m.get("margin_percent") or "").strip():
				add(rec, "margin-percent", "margin row without a Margin %")

	return issues, resolved, gst_dupes, cust_conflicts


def _effective_severity(
	issue,
	create_masters,
	clear_missing_links,
	allow_duplicate_gst,
	update_existing=True,
	renumber_collisions=False,
):
	"""A blocking issue drops to 'auto' once the flag that handles it is on."""
	cat = issue["category"]
	if cat in ("missing-state", "missing-city") and create_masters:
		return "auto"
	if cat == "missing-link" and clear_missing_links:
		return "auto"
	if cat == "duplicate-gstin" and allow_duplicate_gst:
		return "auto"
	if cat == "id-collision" and (renumber_collisions or not update_existing):
		return "auto"  # renumbered or skipped, nothing is overwritten
	return issue["severity"]


def _apply_skips(records, skip_records):
	"""Drop records the caller asked to leave out, by Buyer Master ID."""
	if not skip_records:
		return records
	if isinstance(skip_records, str):
		skip = {s.strip() for s in skip_records.split(",") if s.strip()}
	else:
		skip = {str(s).strip() for s in skip_records}
	kept = [r for r in records if r["name"] not in skip]
	dropped = len(records) - len(kept)
	if dropped:
		print(f"  skipping {dropped} record(s) by request: {sorted(skip)}")
	missing = skip - {r["name"] for r in records}
	if missing:
		print(f"  (skip_records not found in the sheet: {sorted(missing)})")
	return kept


def _sibling(path, suffix):
	return os.path.join(
		os.path.dirname(os.path.abspath(path)),
		os.path.splitext(os.path.basename(path))[0] + suffix,
	)


def preflight(
	path=None,
	limit=None,
	create_masters=False,
	clear_missing_links=False,
	allow_duplicate_gst=False,
	update_existing=True,
	renumber_collisions=False,
	skip_records=None,
	report_path=None,
):
	"""Report everything that would block the import. Writes nothing to the DB."""
	records = _apply_skips(parse(path), skip_records)
	if limit:
		records = records[: int(limit)]
	issues, resolved, _, _ = analyse(records)

	for i in issues:
		i["severity"] = _effective_severity(
			i, create_masters, clear_missing_links, allow_duplicate_gst, update_existing, renumber_collisions
		)
	blocking = [i for i in issues if i["severity"] == "blocking"]

	report_path = report_path or _sibling(path, "_preflight.csv")
	with open(report_path, "w", newline="", encoding="utf-8") as f:
		w = csv.writer(f)
		w.writerow(["Buyer Master", "Sheet line", "Severity", "Category", "Message", "Value"])
		for i in issues:
			w.writerow([i["name"], i["line"], i["severity"], i["category"], i["message"], i["value"]])

	print(f"\n=== Preflight: {os.path.basename(path)} ===")
	print(f"  records                {len(records)}")
	print(f"  already in the DB      {len(resolved[DOCTYPE])}")
	print(f"  address rows           {sum(len(r['addresses']) for r in records)}")
	print(f"  margin rows            {sum(len(r['margins']) for r in records)}")
	print(f"  records blocked        {len({i['name'] for i in blocking})}")

	counts = Counter((i["severity"], i["category"]) for i in issues)
	if counts:
		print("\n  Issues by category:")
		for (sev, cat), n in sorted(counts.items(), key=lambda kv: (kv[0][0] != "blocking", -kv[1])):
			distinct = len({i["value"] for i in issues if i["category"] == cat and i["value"]})
			extra = f", {distinct} distinct value(s)" if distinct else ""
			print(f"    {sev:9} {cat:20} {n:5} row(s){extra}   [{SEVERITY.get(cat, ('', '?'))[1]}]")

	for cat in ("missing-state", "missing-city", "missing-link", "missing-sku"):
		vals = sorted({i["value"] for i in issues if i["category"] == cat and i["value"]})
		if vals:
			print(f"\n  {cat} ({len(vals)}): {vals[:15]}{' ...' if len(vals) > 15 else ''}")

	if blocking:
		print(f"\n  {len(blocking)} blocking issue(s) across {len({i['name'] for i in blocking})} record(s):")
		for i in blocking[:20]:
			print(f"    line {i['line']:5} {i['name']:16} {i['category']:18} {i['message'][:80]}")
		if len(blocking) > 20:
			print(f"    ... {len(blocking) - 20} more")
	else:
		print("\n  No blocking issues — safe to run().")
	print(f"\n  full report: {report_path}")

	return {"records": len(records), "issues": len(issues), "blocking": len(blocking), "report": report_path}


# ---------------------------------------------------------------- import


class _Cache:
	"""Remembers link lookups (and rows this run created) to keep the DB quiet."""

	def __init__(self, resolved=None):
		self._seen = {}
		for dt, names in (resolved or {}).items():
			for n in names:
				self._seen[(dt, n)] = True

	def exists(self, doctype, name):
		key = (doctype, name)
		if key not in self._seen:
			self._seen[key] = bool(frappe.db.exists(doctype, name))
		return self._seen[key]

	def add(self, doctype, name):
		self._seen[(doctype, name)] = True


def _ensure_state(cache, state, country, notes):
	if not state or cache.exists("State", state):
		return bool(state)
	if not country or not cache.exists("Country", country):
		country = "India"
	doc = frappe.new_doc("State")
	doc.state_name = state
	doc.country = country
	doc.insert(ignore_permissions=True)
	cache.add("State", doc.name)
	notes.append(f"created State '{doc.name}'")
	return True


def _ensure_city(cache, city, state, country, notes):
	if not city or cache.exists("City", city):
		return bool(city)
	if not state or not cache.exists("State", state):
		return False  # a City needs a valid State; the caller clears the value
	if not country or not cache.exists("Country", country):
		country = "India"
	doc = frappe.new_doc("City")
	doc.city_name = city
	doc.state = state
	doc.country = country
	doc.insert(ignore_permissions=True)
	cache.add("City", doc.name)
	notes.append(f"created City '{doc.name}'")
	return True


def _ensure_masters(records, cache):
	"""Create every State/City the sheet needs, as a committed pre-pass (outside the import loop, so a record's rollback can't drop cached masters)."""
	made = []
	for rec in records:  # states first, a City row needs a valid State
		d = rec["doc"]
		for a in rec["addresses"]:
			_ensure_state(cache, (a.get("state") or "").strip(), (a.get("country") or "").strip(), made)
		country = (d.get("country") or "").strip()
		_ensure_state(cache, (d.get("state") or "").strip(), country, made)
		_ensure_state(cache, (d.get("shipping_state") or "").strip(), country, made)
	states = len(made)
	for rec in records:
		d = rec["doc"]
		for a in rec["addresses"]:
			_ensure_city(
				cache,
				(a.get("city") or "").strip(),
				(a.get("state") or "").strip(),
				(a.get("country") or "").strip(),
				made,
			)
		country = (d.get("country") or "").strip()
		_ensure_city(cache, (d.get("city") or "").strip(), (d.get("state") or "").strip(), country, made)
		_ensure_city(
			cache, (d.get("shipping_city") or "").strip(), (d.get("shipping_state") or "").strip(), country, made
		)
	if made:
		frappe.db.commit()
	return states, len(made) - states


def _ensure_customers(records, cache):
	"""Create every Customer the sheet names, under that exact ID, as a whole-file pre-pass."""
	from alpinos.alpinos_development.doctype.buyer_master.buyer_master import (
		_default_company,
		_selling_defaults,
	)

	cg, territory = _selling_defaults()
	if not cg or not territory:
		frappe.throw("Set Customer Group and Territory in Selling Settings before importing.")
	company = _default_company()

	made = 0
	for rec in records:
		d = rec["doc"]
		cust_id = (d.get("customer") or "").strip()
		biz = (d.get("customer_business_name") or "").strip()
		if not cust_id or not biz or cache.exists("Customer", cust_id):
			continue
		cust = frappe.new_doc("Customer")
		cust.customer_name = biz  # plain name shown on SOs and stickers
		cust.customer_type = "Company"
		cust.customer_group = cg
		cust.territory = territory
		# custom_order_type is a link to Alpino Customer Type; an unknown value fails the insert
		if d.get("customer_type") and cache.exists("Alpino Customer Type", d["customer_type"]):
			cust.custom_order_type = d["customer_type"]
		if d.get("gst_type") == "Registered Business" and d.get("gst_no"):
			cust.tax_id = (d["gst_no"] or "").strip().upper()
		elif d.get("gst_type") == "Unregistered Business" and d.get("pan_no"):
			cust.tax_id = d["pan_no"]
		if company:
			cust.append("companies", {"company": company})
		cust.flags.ignore_mandatory = True
		cust.insert(ignore_permissions=True, set_name=cust_id)
		cache.add("Customer", cust.name)
		made += 1
	if made:
		frappe.db.commit()
	return made


def _sanitise(rec, cache, on_missing_sku, clear_missing_links):
	"""Fix up link values so the save cannot fail on a LinkValidationError."""
	notes = []
	d = rec["doc"]

	for fieldname, dt in PARENT_LINKS.items():
		val = (d.get(fieldname) or "").strip()
		if not val or cache.exists(dt, val):
			continue
		if dt in ("State", "City") or clear_missing_links:
			notes.append(f"cleared {fieldname}='{val}' ({dt} not found)")
			d.pop(fieldname, None)
		# otherwise leave it in place and let the record fail loudly

	for a in rec["addresses"]:
		for fieldname, dt in ADDRESS_LINKS.items():
			val = (a.get(fieldname) or "").strip()
			if val and not cache.exists(dt, val):
				notes.append(f"address row: cleared {fieldname}='{val}'")
				a.pop(fieldname, None)

	kept = []
	for m in rec["margins"]:
		sku = (m.get("sku") or "").strip()
		if sku and not cache.exists("Item", sku):
			if on_missing_sku == "blank":
				notes.append(f"margin: blanked missing SKU '{sku}'")
				m.pop("sku", None)
			else:
				notes.append(f"margin: skipped row for missing Item '{sku}'")
				continue
		ig = (m.get("item_group") or "").strip()
		if ig and not cache.exists("Item Group", ig):
			notes.append(f"margin: cleared item_group='{ig}'")
			m.pop("item_group", None)
		kept.append(m)
	rec["margins"] = kept
	return notes


def _bump_series(names):
	"""Keep autoname ahead of the IDs we forced in, so new buyers do not collide."""
	highest = {}
	for name in names:
		m = re.match(r"^(.*?-)(\d+)$", name or "")
		if not m:
			continue
		prefix, num = m.group(1), int(m.group(2))
		highest[prefix] = max(highest.get(prefix, 0), num)
	for prefix, num in highest.items():
		frappe.db.sql(
			"""insert into `tabSeries` (name, current) values (%s, %s)
			on duplicate key update current = greatest(current, %s)""",
			(prefix, num, num),
		)
	return highest


def run(
	path=None,
	update_existing=True,
	create_masters=False,
	clear_missing_links=False,
	allow_duplicate_gst=False,
	renumber_collisions=False,
	skip_records=None,
	on_missing_sku="skip",
	force=False,
	limit=None,
	report_path=None,
):
	if not path:
		frappe.throw("path is required")
	if not os.path.exists(path):
		frappe.throw(f"File not found: {path}")
	if on_missing_sku not in ("skip", "blank"):
		frappe.throw("on_missing_sku must be 'skip' or 'blank'")

	records = _apply_skips(parse(path), skip_records)
	if limit:
		records = records[: int(limit)]
	issues, resolved, gst_dupes, cust_conflicts = analyse(records)
	blocking = [
		i
		for i in issues
		if _effective_severity(
			i, create_masters, clear_missing_links, allow_duplicate_gst, update_existing, renumber_collisions
		)
		== "blocking"
	]
	renumber = (
		{i["name"] for i in issues if i["category"] == "id-collision"} if renumber_collisions else set()
	)

	if blocking and not force:
		names = {i["name"] for i in blocking}
		print(f"\n=== {len(blocking)} blocking issue(s) across {len(names)} record(s) — nothing imported ===")
		for (cat, fix), n in Counter(
			(i["category"], SEVERITY.get(i["category"], ("", "?"))[1]) for i in blocking
		).most_common():
			print(f"  {cat:20} {n:5}  -> {fix}")
		print("\nRun preflight for the full list, then fix the sheet or pass the flags above.")
		print("force=True imports everything else and lets the bad records fail individually.")
		return {"aborted": True, "blocking": len(blocking), "records": len(records)}

	records = _order_parents_first(records)
	cache = _Cache(resolved)
	if not allow_duplicate_gst:
		gst_dupes = set()

	# two rows naming one Customer are handed back to the controller to re-derive;
	# done before the pre-pass so the contested ID is never created for the wrong buyer
	for rec in records:
		if rec["name"] in cust_conflicts and rec["doc"].get("customer"):
			rec["dropped_customer"] = rec["doc"].pop("customer")

	if create_masters:
		new_states, new_cities = _ensure_masters(records, cache)
		print(f"\n=== Pre-created {new_states} State(s) and {new_cities} City(ies) ===")
	made = _ensure_customers(records, cache)
	print(f"=== Pre-created {made} Customer(s) under the sheet's own IDs ===")
	print(f"=== Importing {len(records)} Buyer Master record(s) from {os.path.basename(path)} ===")

	results = []
	pending_gst = []
	created = updated = skipped = failed = 0

	for i, rec in enumerate(records, start=1):
		name = rec["name"]
		try:
			notes = _sanitise(rec, cache, on_missing_sku, clear_missing_links)
			if rec.get("dropped_customer"):
				notes.append(f"customer '{rec['dropped_customer']}' is taken — the controller derives a site-scoped ID")
			# sheet ID collides with a different buyer here; insert under a fresh ID rather than overwrite
			renumbered = name in renumber
			exists = frappe.db.exists(DOCTYPE, name) and not renumbered
			if exists and not update_existing:
				results.append([name, rec["line"], "skipped", "already exists", ""])
				skipped += 1
				continue

			# GSTIN is held by an unrelated buyer too; save without it and write it after the run.
			# validate() tolerates the unchanged value; writing it now would break later buyers.
			deferred_gst = None
			if name in gst_dupes and rec["doc"].get("gst_no"):
				deferred_gst = (rec["doc"].pop("gst_no") or "").strip().upper()
				notes.append(f"GST {deferred_gst} written after the run (already held by an unrelated buyer)")

			if exists:
				doc = frappe.get_doc(DOCTYPE, name)
				doc.update(rec["doc"])
				doc.set("addresses", [])
				doc.set("margins", [])
			else:
				doc = frappe.new_doc(DOCTYPE)
				doc.update(rec["doc"])
				if not doc.get("customer_id") and not renumbered:
					doc.customer_id = name  # what before_insert intends, but name is unset there

			for a in rec["addresses"]:
				doc.append("addresses", a)
			for m in rec["margins"]:
				doc.append("margins", m)

			if exists:
				doc.save(ignore_permissions=True)
				updated += 1
				action = "updated"
			elif renumbered:
				doc.insert(ignore_permissions=True)  # let autoname pick a free ID
				frappe.db.set_value(DOCTYPE, doc.name, "customer_id", doc.name, update_modified=False)
				created += 1
				action = "created"
				notes.append(f"sheet ID {name} belongs to another buyer here — inserted as {doc.name}")
			else:
				# set_name forces the sheet's own ID, so parent_buyer resolves
				doc.insert(ignore_permissions=True, set_name=name)
				created += 1
				action = "created"

			if deferred_gst:
				pending_gst.append((doc.name, doc.customer, deferred_gst))

			# controller may rename an adopted Customer once its ID is free; note it in the report
			wanted_customer = (rec["doc"].get("customer") or "").strip()
			if wanted_customer and doc.customer != wanted_customer:
				notes.append(f"customer renamed by the controller: '{wanted_customer}' -> '{doc.customer}'")

			frappe.db.commit()
			cache.add(DOCTYPE, doc.name)
			results.append([name, rec["line"], action, "", "; ".join(notes)])
		except Exception as e:
			frappe.db.rollback()
			failed += 1
			msg = " ".join(str(e).split())[:400]
			# after_insert commits, so a later failure can leave the row behind
			partial = "row exists despite the failure — review it" if frappe.db.exists(DOCTYPE, name) else ""
			results.append([name, rec["line"], "failed", msg, partial])
			print(f"  [{i}/{len(records)}] {name} FAILED: {msg[:160]}")

		if i % 50 == 0:
			print(f"  [{i}/{len(records)}] created={created} updated={updated} failed={failed}")

	for bm_name, customer, gst in pending_gst:
		frappe.db.set_value(DOCTYPE, bm_name, "gst_no", gst, update_modified=False)
		if customer:
			frappe.db.set_value("Customer", customer, "tax_id", gst, update_modified=False)
	if pending_gst:
		print(f"  wrote {len(pending_gst)} deferred GSTIN(s) held by more than one buyer")

	series = _bump_series([r["name"] for r in records])
	frappe.db.commit()

	report_path = report_path or _sibling(path, "_import_result.csv")
	with open(report_path, "w", newline="", encoding="utf-8") as f:
		w = csv.writer(f)
		w.writerow(["Buyer Master", "Sheet line", "Result", "Error", "Adjustments"])
		w.writerows(results)

	print(
		f"\n=== Done: created={created} updated={updated} skipped={skipped} failed={failed} ==="
		f"\n  naming series bumped to: {series}"
		f"\n  report: {report_path}"
	)
	return {
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"failed": failed,
		"report": report_path,
	}


def verify(path=None, limit=None, report_path=None):
	"""Compare the site against the sheet after an import (margin rows with an unknown SKU counted separately)."""
	records = parse(path)
	if limit:
		records = records[: int(limit)]
	_, resolved, _, _ = analyse(records)
	items = _lower(resolved.get("Item", set()))

	missing, mismatched, rows = [], [], []
	for rec in records:
		name, d = rec["name"], rec["doc"]
		if not frappe.db.exists(DOCTYPE, name):
			missing.append(name)
			rows.append([name, rec["line"], "not imported", ""])
			continue
		doc = frappe.get_doc(DOCTYPE, name)
		want_margins = [m for m in rec["margins"] if not m.get("sku") or (m["sku"] or "").lower() in items]
		problems = []
		if (doc.customer_business_name or "").strip() != (d.get("customer_business_name") or "").strip():
			problems.append(f"business name '{doc.customer_business_name}' != '{d.get('customer_business_name')}'")
		if (doc.gst_no or "").strip().upper() != (d.get("gst_no") or "").strip().upper():
			problems.append(f"gst '{doc.gst_no}' != '{d.get('gst_no')}'")
		if (doc.parent_buyer or "") != (d.get("parent_buyer") or ""):
			problems.append(f"parent '{doc.parent_buyer}' != '{d.get('parent_buyer')}'")
		if len(doc.addresses) != len(rec["addresses"]):
			problems.append(f"{len(doc.addresses)} address rows != {len(rec['addresses'])}")
		if len(doc.margins) != len(want_margins):
			problems.append(f"{len(doc.margins)} margin rows != {len(want_margins)} importable")
		if doc.customer and not frappe.db.exists("Customer", doc.customer):
			problems.append(f"customer '{doc.customer}' does not exist")
		if problems:
			mismatched.append(name)
			rows.append([name, rec["line"], "mismatch", "; ".join(problems)])

	report_path = report_path or _sibling(path, "_verify.csv")
	with open(report_path, "w", newline="", encoding="utf-8") as f:
		w = csv.writer(f)
		w.writerow(["Buyer Master", "Sheet line", "Result", "Detail"])
		w.writerows(rows)

	print(f"\n=== Verify: {os.path.basename(path)} ===")
	print(f"  records in the sheet   {len(records)}")
	print(f"  matching the sheet     {len(records) - len(missing) - len(mismatched)}")
	print(f"  not imported           {len(missing)}")
	print(f"  imported but differing {len(mismatched)}")
	skipped_margins = sum(
		1 for r in records for m in r["margins"] if m.get("sku") and (m["sku"] or "").lower() not in items
	)
	print(f"  margin rows dropped for a missing Item: {skipped_margins}")
	for row in rows[:15]:
		print(f"    {row[0]:16} line {row[1]:5} {row[2]:12} {row[3][:70]}")
	if len(rows) > 15:
		print(f"    ... {len(rows) - 15} more")
	print(f"\n  full report: {report_path}")
	return {"records": len(records), "missing": len(missing), "mismatched": len(mismatched), "report": report_path}
