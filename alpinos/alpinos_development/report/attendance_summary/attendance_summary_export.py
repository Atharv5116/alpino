# Copyright (c) 2026, Alpinos and contributors
# License: MIT
"""Final-Format Excel export (Summary + Details sheets) for the Attendance Summary report."""

import io

import frappe
from frappe import _

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from alpinos.alpinos_development.report.attendance_summary import attendance_summary as rpt

# colours (mirror attendance_summary.js)
RED = "FFC00000"
BLACK = "FF000000"
WHITE = "FFFFFFFF"

# Summary section bands by fieldname
SECTION = {
	"employee_name": "basic", "employee": "basic", "status": "basic", "date_of_joining": "basic",
	"aging": "basic", "department": "basic", "company": "basic",
	"month_working_days": "days", "final_paid_days": "days", "final_payable_days": "days",
	"present_working_days": "days", "clock_in_days": "days",
	"absent_days": "ded", "late_entries": "ded", "late_half_days": "ded", "late_full_days": "ded",
	"working_hours_shortage": "ded",
	"paid_leave": "leave", "unpaid_leave": "leave", "wfh": "leave", "od": "leave",
	"public_holiday": "other", "weekend": "other", "missing_attendance": "other", "avg_working_hours": "other",
	"verify": "verify",
}
CELL_BG = {"basic": "FFEEF2FF", "days": "FFE8F6EE", "ded": "FFFDECEA", "leave": "FFFFF8E1", "other": "FFF4E9FD", "verify": "FFECEFF1"}
HEAD_BG = {"basic": "FFC7D2FE", "days": "FFB7E4C7", "ded": "FFF6BDB6", "leave": "FFFFE39A", "other": "FFE0C3FB", "verify": "FFCFD8DC"}

# Details cell fills.
LEAVE_FILL = "FFCCC0DA"      # leave day (purple, matches the spec sheet)
WEEKEND_FILL = "FFEDEDED"    # weekly-off
HOLIDAY_COLOR = "FF1976D2"   # holiday text (blue)
EMP_HEAD_FILL = "FF4472C4"   # "Employee Name" header (blue)
DATE_BANNER = "FFC00000"     # "Date" banner text (red)

_thin = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


@frappe.whitelist()
def download_final_format(month, employee=None, company=None):
	"""Build the Final-Format workbook for the given filters and stream it."""
	if not frappe.has_permission("Attendance", "read"):
		frappe.throw(_("Not permitted"), frappe.PermissionError)
	filters = {"month": month}
	if employee:
		filters["employee"] = employee
	if company:
		filters["company"] = company

	columns, data = rpt.execute(filters)
	data = [frappe._dict(r) for r in (data or [])]

	wb = Workbook()
	_build_summary(wb.active, columns, data)
	_build_details(wb.create_sheet("Details"), columns, data)

	bio = io.BytesIO()
	wb.save(bio)
	frappe.response["filename"] = f"Attendance_Summary_{month}.xlsx"
	frappe.response["filecontent"] = bio.getvalue()
	frappe.response["type"] = "binary"


def _build_summary(ws, columns, data):
	ws.title = "Summary"
	cols = [c for c in columns if not c["fieldname"].startswith("day_")]

	for j, c in enumerate(cols, start=1):
		cell = ws.cell(row=1, column=j, value=c["label"])
		sec = SECTION.get(c["fieldname"])
		cell.font = Font(bold=True, color=BLACK)
		cell.alignment = CENTER
		cell.border = BORDER
		if sec:
			cell.fill = PatternFill("solid", fgColor=HEAD_BG[sec])
		ws.column_dimensions[get_column_letter(j)].width = max(10, min(46, (c.get("width") or 110) / 6.5))

	for i, row in enumerate(data, start=2):
		for j, c in enumerate(cols, start=1):
			val = row.get(c["fieldname"])
			cell = ws.cell(row=i, column=j, value="" if val is None else val)
			cell.border = BORDER
			cell.alignment = Alignment(vertical="top", wrap_text=True)
			sec = SECTION.get(c["fieldname"])
			if sec:
				cell.fill = PatternFill("solid", fgColor=CELL_BG[sec])

	ws.freeze_panes = "C2"   # keep Employee Name + ID and the header visible


def _build_details(ws, columns, data):
	day_cols = [c for c in columns if c["fieldname"].startswith("day_")]
	n_days = len(day_cols)

	# Row 1: "Employee Name" (A1:A2, blue) + "Date" banner across the day columns.
	ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
	a1 = ws.cell(row=1, column=1, value="Employee Name")
	a1.fill = PatternFill("solid", fgColor=EMP_HEAD_FILL)
	a1.font = Font(bold=True, color=WHITE)
	a1.alignment = CENTER
	a1.border = BORDER
	if n_days:
		ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1 + n_days)
		banner = ws.cell(row=1, column=2, value="Date")
		banner.font = Font(bold=True, color=DATE_BANNER, size=12)
		banner.alignment = CENTER
		banner.border = BORDER

	# Row 2: day headers ("1 Wed", "2 Thu", ...).
	for j, c in enumerate(day_cols, start=2):
		cell = ws.cell(row=2, column=j, value=c["label"])
		cell.font = Font(bold=True, color=BLACK)
		cell.alignment = Alignment(horizontal="left", vertical="center")
		cell.border = BORDER
		ws.column_dimensions[get_column_letter(j)].width = 30

	ws.column_dimensions["A"].width = 22
	ws.row_dimensions[2].height = 18

	# Data rows.
	for i, row in enumerate(data, start=3):
		name = ws.cell(row=i, column=1, value=row.get("employee_name") or "")
		name.alignment = Alignment(vertical="top", wrap_text=True)
		name.border = BORDER
		name.font = Font(color=BLACK)
		for j, c in enumerate(day_cols, start=2):
			_write_day_cell(ws.cell(row=i, column=j), str(row.get(c["fieldname"]) or ""))
		ws.row_dimensions[i].height = 108

	ws.freeze_panes = "B3"   # keep Employee Name column + the two header rows


def _write_day_cell(cell, text):
	"""Render one per-day cell: rich attendance text, or a shaded leave / weekend cell."""
	cell.border = BORDER
	cell.alignment = WRAP_TOP
	t = (text or "").strip()
	if t in ("", "-"):
		return
	if "In:" in t:  # attendance detail cell
		cell.value = _attendance_rich(t)
		return
	if t.startswith("WEEKEND"):
		cell.value = t
		cell.font = Font(bold=True, color="FF607D8B")
		cell.fill = PatternFill("solid", fgColor=WEEKEND_FILL)
		cell.alignment = CENTER
		return
	if t.startswith("HOLIDAY"):
		cell.value = t
		cell.font = Font(color=HOLIDAY_COLOR)
		cell.alignment = CENTER
		return
	# Leave day (leave type / "HALF DAY - ...").
	cell.value = t
	cell.font = Font(bold=True, color=BLACK)
	cell.fill = PatternFill("solid", fgColor=LEAVE_FILL)
	cell.alignment = CENTER


def _attendance_rich(text):
	"""Multi-line attendance cell as rich text; whole cell red when ABSENT."""
	lines = text.split("\n")
	absent = any(ln.startswith("ABSENT") for ln in lines)
	label_font = InlineFont(b=True, color=RED if absent else BLACK)
	value_font = InlineFont(color=RED if absent else BLACK)
	red_font = InlineFont(b=True, color=RED)

	# Carry the line break inside a TextBlock's text: a bare "\n" element in a CellRichText
	# is written without xml:space="preserve", so Excel strips it and the lines run together.
	parts = []
	for i, line in enumerate(lines):
		nl = "\n" if i else ""
		if line in ("WFH", "OD") or line.startswith("ABSENT"):
			parts.append(TextBlock(red_font, nl + line))
			continue
		idx = line.find(":")
		if idx == -1:
			parts.append(TextBlock(value_font, nl + line))
			continue
		label, value = line[: idx + 1], line[idx + 1:]
		parts.append(TextBlock(label_font, nl + label))
		lbl = label.strip().lower()
		if value.strip() and (lbl.startswith("late time") or lbl.startswith("early out")) and not absent:
			parts.append(TextBlock(red_font, value))
		elif value:
			parts.append(TextBlock(value_font, value))
	return CellRichText(parts)
